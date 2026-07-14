"""Load the per-month tables out of an .xlsx workbook.

Two layouts are supported, in this order of preference:

1. Real Excel Tables (ListObjects) named ``tbl_Jan`` ... ``tbl_Dec`` -- the layout the
   Power Query template expects, since ``Excel.CurrentWorkbook()`` only sees named tables.
2. Plain worksheets titled ``Jan`` ... ``Dec``, with a header row in row 1. This is the
   fallback for the far more common workbook that never had tables defined.

Months that are absent are simply not returned; the diff engine treats a gap the same
way the M code does.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

MONTHS: List[str] = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

DEFAULT_TABLE_PREFIX = "tbl_"
DEFAULT_KEY = "ID"


class WorkbookError(Exception):
    """The workbook cannot be diffed as-is (missing file, no months, missing key)."""


@dataclass
class MonthTable:
    """One month's rows, keyed by the diff key."""

    month: str
    columns: List[str]
    rows: List[Dict[str, Any]] = field(default_factory=list)
    source: str = ""

    def by_key(self, key: str) -> "Dict[Any, Dict[str, Any]]":
        """Index the rows by key. Later duplicates win, matching a join's last-write."""
        return {row[key]: row for row in self.rows if row.get(key) is not None}


def _clean(value: Any) -> Any:
    """Normalize a cell value: blank-ish strings become None, others are stripped."""
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _rows_from_grid(grid: Sequence[Sequence[Any]]) -> "tuple[List[str], List[Dict[str, Any]]]":
    """Turn a rectangular block of cell values (header row first) into typed rows."""
    if not grid:
        return [], []

    header_cells = [_clean(cell) for cell in grid[0]]
    # Trailing unnamed columns are Excel padding, not data.
    while header_cells and header_cells[-1] is None:
        header_cells.pop()
    if not header_cells:
        return [], []

    columns = [str(cell) if cell is not None else f"Column{i + 1}" for i, cell in enumerate(header_cells)]

    rows: List[Dict[str, Any]] = []
    for raw in grid[1:]:
        values = [_clean(cell) for cell in raw[: len(columns)]]
        if all(value is None for value in values):
            continue  # spacer row
        values += [None] * (len(columns) - len(values))
        rows.append(dict(zip(columns, values)))

    return columns, rows


def _read_named_tables(workbook, prefix: str, months: Sequence[str]) -> Dict[str, MonthTable]:
    wanted = {f"{prefix}{month}".casefold(): month for month in months}
    found: Dict[str, MonthTable] = {}

    for worksheet in workbook.worksheets:
        # ws.tables is a TableList: .items() yields (name, ref) pairs, .values() the tables.
        for table in worksheet.tables.values():
            month = wanted.get(str(table.name).casefold())
            if month is None:
                continue

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            grid = [
                [cell.value for cell in row]
                for row in worksheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                )
            ]
            columns, rows = _rows_from_grid(grid)
            if columns:
                found[month] = MonthTable(month, columns, rows, source=f"table {table.name}")

    return found


def _read_sheets(workbook, months: Sequence[str]) -> Dict[str, MonthTable]:
    wanted = {month.casefold(): month for month in months}
    found: Dict[str, MonthTable] = {}

    for worksheet in workbook.worksheets:
        month = wanted.get(str(worksheet.title).strip().casefold())
        if month is None:
            continue

        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        columns, rows = _rows_from_grid(grid)
        if columns:
            found[month] = MonthTable(month, columns, rows, source=f"sheet {worksheet.title}")

    return found


def read_workbook(
    path: "str | Path",
    key: str = DEFAULT_KEY,
    table_prefix: str = DEFAULT_TABLE_PREFIX,
    months: Optional[Sequence[str]] = None,
) -> Dict[str, MonthTable]:
    """Read every month present in the workbook, keyed by month name.

    Raises WorkbookError if the file is unreadable, no month is found, or a month that
    was found lacks the key column -- the diff is meaningless without a join key.
    """
    path = Path(path)
    if not path.is_file():
        raise WorkbookError(f"No such workbook: {path}")

    months = list(months) if months else MONTHS

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # openpyxl raises a grab-bag of errors on bad input
        raise WorkbookError(f"Could not open {path.name} as an .xlsx workbook: {exc}") from exc

    try:
        tables = _read_named_tables(workbook, table_prefix, months)
        if not tables:
            tables = _read_sheets(workbook, months)
    finally:
        workbook.close()

    if not tables:
        raise WorkbookError(
            f"{path.name} has no month tables. Expected Excel tables named "
            f"'{table_prefix}Jan'...'{table_prefix}Dec', or sheets named 'Jan'...'Dec'."
        )

    missing_key = sorted(
        (table.month for table in tables.values() if key not in table.columns),
        key=months.index,
    )
    if missing_key:
        raise WorkbookError(
            f"Key column {key!r} is missing from: {', '.join(missing_key)}. "
            f"Pass --key to name the column that identifies a row."
        )

    return {month: tables[month] for month in months if month in tables}
