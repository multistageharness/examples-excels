"""Render a DiffResult as a console table, CSV, JSON, or a styled .xlsx sheet."""

from __future__ import annotations

import csv
import datetime as dt
import io
import json
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .diff import ADDED, MODIFIED, REMOVED, DiffResult

FORMATS = ("table", "csv", "json", "xlsx")

_HEADER_FILL = PatternFill("solid", fgColor="ED7D31")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_STATUS_FILLS = {
    ADDED: PatternFill("solid", fgColor="E2EFDA"),
    REMOVED: PatternFill("solid", fgColor="FCE4E4"),
    MODIFIED: PatternFill("solid", fgColor="FFF2CC"),
}


def format_for(path: "str | Path | None", requested: "str | None") -> str:
    """Pick the output format: an explicit --format wins, else infer from the extension."""
    if requested:
        return requested
    if path is not None:
        suffix = Path(path).suffix.casefold()
        if suffix in (".xlsx", ".xlsm"):
            return "xlsx"
        if suffix == ".json":
            return "json"
        if suffix in (".csv", ".txt"):
            return "csv"
    return "table"


def _scalar(value: Any) -> Any:
    """Excel dates arrive as datetimes; make them JSON- and CSV-safe."""
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def _text(value: Any) -> str:
    return "" if value is None else str(_scalar(value))


def to_table(result: DiffResult) -> str:
    """A fixed-width table for the terminal."""
    header = result.header
    rows = [[_text(row[column]) for column in header] for row in result.as_dicts()]

    widths = [len(column) for column in header]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))

    def line(cells: List[str]) -> str:
        return "  ".join(cell.ljust(widths[i]) for i, cell in enumerate(cells)).rstrip()

    out = [line(header), "  ".join("-" * width for width in widths)]
    out.extend(line(row) for row in rows)
    return "\n".join(out)


def to_csv(result: DiffResult) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=result.header, extrasaction="ignore")
    writer.writeheader()
    for row in result.as_dicts():
        writer.writerow({column: _scalar(value) for column, value in row.items()})
    return buffer.getvalue()


def to_json(result: DiffResult) -> str:
    payload = [
        {column: _scalar(value) for column, value in row.items()} for row in result.as_dicts()
    ]
    return json.dumps(payload, indent=2, default=str)


def write_xlsx(result: DiffResult, path: "str | Path") -> None:
    """Write the changes to a single sheet, styled like the Power Query output."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Changes"

    header = result.header
    rows = result.as_dicts()

    worksheet.append(header)
    for row in rows:
        worksheet.append([_scalar(row[column]) for column in header])

    status_index = header.index("Change_Status")

    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        fill = _STATUS_FILLS.get(row[status_index].value)
        if fill is not None:
            for cell in row:
                cell.fill = fill

    for index, column in enumerate(header, start=1):
        longest = max([len(column)] + [len(_text(row[column])) for row in rows])
        worksheet.column_dimensions[get_column_letter(index)].width = min(longest + 4, 40)

    # An Excel Table needs at least one data row; below that, fall back to an autofilter.
    span = f"A1:{get_column_letter(len(header))}{worksheet.max_row}"
    if worksheet.max_row > 1:
        table = Table(displayName="tbl_Changes", ref=span)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = span

    worksheet.freeze_panes = "A2"
    workbook.save(Path(path))


def render(result: DiffResult, fmt: str) -> str:
    """Render to a string. ``xlsx`` is binary and must go through write_xlsx instead."""
    if fmt == "table":
        return to_table(result)
    if fmt == "csv":
        return to_csv(result)
    if fmt == "json":
        return to_json(result)
    raise ValueError(f"{fmt!r} cannot be rendered as text; use write_xlsx()")
