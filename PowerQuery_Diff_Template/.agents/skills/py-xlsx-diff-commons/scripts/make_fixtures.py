#!/usr/bin/env python3
"""Generate the fixture workbooks this skill suite is tested against.

    python3 make_fixtures.py [OUTDIR]      # default: ./fixtures

Fixtures are generated, never committed as binary blobs, so they can be regenerated on
demand and diffed as code.

    sample.xlsx   the canonical 3-month workbook, as real Excel Tables (tbl_Jan..tbl_Mar).
                  Byte-for-byte the same shape as the suite's reference sample.xlsx, and
                  produces the reference output documented in reference/CONTRACT.md.
    sheets.xlsx   the same data as plain sheets named Jan/Feb/Mar and NO named tables --
                  exercises the sheet-fallback layout (REQ-2.1.2) that Power Query cannot read.
    gap.xlsx      Jan and Mar only. Feb is absent, so Mar's predecessor is missing and Mar
                  is itself a base month -- the List.Generate short-circuit (REQ-3.5.1).
    messy.xlsx    the edge cases the reader normalizes: a duplicate key, a null key, a spacer
                  row, a trailing unnamed column, and a column that exists in only one month.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMNS = ["ID", "Status", "Owner"]

# Bob (102) is Active in Jan, flips to Inactive in Feb, and is gone in Mar -- so his Removed
# row must carry FEB's values, not Jan's. That single row is what the "with_name" variant is for.
MONTHS = {
    "Jan": [
        [101, "Active", "Alice"],
        [102, "Active", "Bob"],
        [103, "Active", "Charlie"],
        [104, "Active", "Dana"],
    ],
    "Feb": [
        [101, "Active", "Alice"],
        [102, "Inactive", "Bob"],
        [104, "Active", "Dana"],
        [105, "Active", "Eve"],
    ],
    "Mar": [
        [101, "Active", "Alice"],
        [104, "Pending", "Dana"],
        [105, "Active", "Eve"],
        [106, "New", "Frank"],
    ],
}


def _add_table(worksheet, month: str, rows: list, columns: list) -> None:
    letter = chr(ord("A") + len(columns) - 1)
    table = Table(displayName=f"tbl_{month}", ref=f"A1:{letter}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    worksheet.add_table(table)


def _sheet(workbook: Workbook, month: str, columns: list, rows: list):
    worksheet = workbook.create_sheet(month)
    worksheet.append(columns)
    for row in rows:
        worksheet.append(row)
    return worksheet


def build_sample(path: Path) -> Path:
    """The canonical layout: one real Excel Table per month."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for month, rows in MONTHS.items():
        worksheet = _sheet(workbook, month, COLUMNS, rows)
        _add_table(worksheet, month, rows, COLUMNS)
        for column, width in zip("ABC", (8, 12, 14)):
            worksheet.column_dimensions[column].width = width

    workbook.save(path)
    return path


def build_sheets(path: Path) -> Path:
    """The fallback layout: sheets named Jan/Feb/Mar, no named tables at all."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for month, rows in MONTHS.items():
        _sheet(workbook, month, COLUMNS, rows)

    workbook.save(path)
    return path


def build_gap(path: Path) -> Path:
    """Jan and Mar, no Feb: Mar's predecessor is missing, so Mar is a base month."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for month in ("Jan", "Mar"):
        worksheet = _sheet(workbook, month, COLUMNS, MONTHS[month])
        _add_table(worksheet, month, MONTHS[month], COLUMNS)

    workbook.save(path)
    return path


def build_messy(path: Path) -> Path:
    """Every normalization rule the reader implements, in one workbook."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    # A trailing unnamed header column (Excel padding) must be dropped, not read as data.
    jan_columns = ["ID", "Status", "Owner", None]
    jan_rows = [
        [201, "Active", "  Alice  ", None],  # whitespace is stripped
        [202, "Active", "Bob", None],
        [None, None, None, None],            # spacer row -- skipped entirely
        [203, "  ", "Charlie", None],        # whitespace-only cell becomes null
        [None, "Active", "Nokey", None],     # null key -- excluded from the join
        [202, "Duplicate", "Bob II", None],  # duplicate key -- the LAST row wins
    ]

    # Feb adds a column Jan never had; it must still appear in the output's column union.
    feb_columns = ["ID", "Status", "Owner", "Region"]
    feb_rows = [
        [201, "Active", "Alice", "East"],
        [204, "New", "Dana", "West"],        # Added
    ]

    for month, columns, rows in (("Jan", jan_columns, jan_rows), ("Feb", feb_columns, feb_rows)):
        worksheet = workbook.create_sheet(month)
        worksheet.append(columns)
        for row in rows:
            worksheet.append(row)

    workbook.save(path)
    return path


BUILDERS = {
    "sample.xlsx": build_sample,
    "sheets.xlsx": build_sheets,
    "gap.xlsx": build_gap,
    "messy.xlsx": build_messy,
}


def main(argv: list) -> int:
    outdir = Path(argv[1]) if len(argv) > 1 else Path("fixtures")
    outdir.mkdir(parents=True, exist_ok=True)

    for name, build in BUILDERS.items():
        print(f"wrote {build(outdir / name)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
