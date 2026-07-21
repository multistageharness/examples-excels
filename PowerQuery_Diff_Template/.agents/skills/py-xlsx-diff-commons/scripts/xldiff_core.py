#!/usr/bin/env python3
"""Self-contained month-over-month diff engine for .xlsx workbooks.

A single-file port of the `xldiff` package (reader + diff + writer + CLI) so every skill in
this suite calls one canonical implementation instead of duplicating the rules. Behavior is
pinned to REQ.md; the requirement id is cited at each decision that has one.

Read it, diff it, write it:

    python3 xldiff_core.py WORKBOOK -o changes.xlsx
    python3 xldiff_core.py WORKBOOK -f json

or import it:

    from xldiff_core import read_workbook, diff_tables, render, write_xlsx
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:  # pragma: no cover - environment guard
    sys.exit("xldiff: openpyxl is required. Install it with: pip install openpyxl")

# --- Contract ---------------------------------------------------------------------------

MONTHS: List[str] = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

DEFAULT_TABLE_PREFIX = "tbl_"  # REQ-2.1.5
DEFAULT_KEY = "ID"             # REQ-3.1.4

ADDED = "Added"
REMOVED = "Removed"
UNCHANGED = "Unchanged"
MODIFIED = "Modified"
BASE_MONTH = "Base Month"

MONTH_COLUMN = "Month"
STATUS_COLUMN = "Change_Status"

#: The statuses that represent an actual change to the roster of rows. REQ-3.2.6
CHANGE_STATUSES = (ADDED, REMOVED, MODIFIED)

FORMATS = ("table", "csv", "json", "xlsx")  # REQ-4.1

_HEADER_FILL = PatternFill("solid", fgColor="ED7D31")   # REQ-4.1.2
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_STATUS_FILLS = {                                       # REQ-4.1.3
    ADDED: PatternFill("solid", fgColor="E2EFDA"),
    REMOVED: PatternFill("solid", fgColor="FCE4E4"),
    MODIFIED: PatternFill("solid", fgColor="FFF2CC"),
}

EXIT_OK = 0
EXIT_ERROR = 1


class WorkbookError(Exception):
    """The workbook cannot be diffed as-is (missing file, no months, missing key). REQ-2.3.6"""


# --- Reader -----------------------------------------------------------------------------


@dataclass
class MonthTable:
    """One month's rows, keyed by the diff key."""

    month: str
    columns: List[str]
    rows: List[Dict[str, Any]] = field(default_factory=list)
    source: str = ""

    def by_key(self, key: str) -> Dict[Any, Dict[str, Any]]:
        """Index the rows by key. Later duplicates win (REQ-3.1.2); null keys drop (REQ-3.1.3)."""
        return {row[key]: row for row in self.rows if row.get(key) is not None}


def _clean(value: Any) -> Any:
    """Blank-ish strings become None, other strings are stripped, types pass through. REQ-2.2.2"""
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _rows_from_grid(grid: Sequence[Sequence[Any]]) -> "tuple[List[str], List[Dict[str, Any]]]":
    """Turn a rectangular block of cell values (header row first) into typed rows."""
    if not grid:
        return [], []

    header_cells = [_clean(cell) for cell in grid[0]]  # REQ-2.2.3
    while header_cells and header_cells[-1] is None:   # REQ-2.2.4 trailing padding
        header_cells.pop()
    if not header_cells:
        return [], []  # REQ-2.2.8

    columns = [
        str(cell) if cell is not None else f"Column{i + 1}"  # REQ-2.2.5
        for i, cell in enumerate(header_cells)
    ]

    rows: List[Dict[str, Any]] = []
    for raw in grid[1:]:
        values = [_clean(cell) for cell in raw[: len(columns)]]  # REQ-2.2.7 truncate
        if all(value is None for value in values):
            continue  # REQ-2.2.6 spacer row
        values += [None] * (len(columns) - len(values))  # REQ-2.2.7 right-pad
        rows.append(dict(zip(columns, values)))

    return columns, rows


def _read_named_tables(workbook, prefix: str, months: Sequence[str]) -> Dict[str, MonthTable]:
    """Layout 1: real Excel Tables named tbl_Jan..tbl_Dec. REQ-2.1.1"""
    wanted = {f"{prefix}{month}".casefold(): month for month in months}  # REQ-2.1.4
    found: Dict[str, MonthTable] = {}

    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            month = wanted.get(str(table.name).casefold())
            if month is None:
                continue

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            grid = [
                [cell.value for cell in row]
                for row in worksheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                )
            ]
            columns, rows = _rows_from_grid(grid)
            if columns:
                found[month] = MonthTable(month, columns, rows, source=f"table {table.name}")

    return found


def _read_sheets(workbook, months: Sequence[str]) -> Dict[str, MonthTable]:
    """Layout 2: plain sheets titled Jan..Dec, header in row 1. REQ-2.1.2"""
    wanted = {month.casefold(): month for month in months}
    found: Dict[str, MonthTable] = {}

    for worksheet in workbook.worksheets:
        month = wanted.get(str(worksheet.title).strip().casefold())  # REQ-2.1.4
        if month is None:
            continue

        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        columns, rows = _rows_from_grid(grid)
        if columns:
            found[month] = MonthTable(month, columns, rows, source=f"sheet {worksheet.title}")

    return found


def read_workbook(
    path: "str | Path",
    key: str = DEFAULT_KEY,
    table_prefix: str = DEFAULT_TABLE_PREFIX,
    months: Optional[Sequence[str]] = None,
) -> Dict[str, MonthTable]:
    """Read every month present in the workbook, keyed by month name.

    Named tables are tried first; the sheet fallback is used only when no table matched at
    all, and the two layouts are never mixed. REQ-2.1.3
    """
    path = Path(path)
    if not path.is_file():
        raise WorkbookError(f"No such workbook: {path}")  # REQ-2.3.1

    months = list(months) if months else MONTHS

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)  # REQ-2.2.1
    except Exception as exc:  # openpyxl raises a grab-bag of errors on bad input
        raise WorkbookError(f"Could not open {path.name} as an .xlsx workbook: {exc}") from exc  # REQ-2.3.2

    try:
        tables = _read_named_tables(workbook, table_prefix, months)
        if not tables:
            tables = _read_sheets(workbook, months)
    finally:
        workbook.close()  # REQ-2.3.7

    if not tables:
        raise WorkbookError(  # REQ-2.3.3
            f"{path.name} has no month tables. Expected Excel tables named "
            f"'{table_prefix}Jan'...'{table_prefix}Dec', or sheets named 'Jan'...'Dec'."
        )

    missing_key = sorted(
        (table.month for table in tables.values() if key not in table.columns),
        key=months.index,
    )
    if missing_key:
        raise WorkbookError(  # REQ-2.3.4
            f"Key column {key!r} is missing from: {', '.join(missing_key)}. "
            f"Pass --key to name the column that identifies a row."
        )

    # Absent months are omitted, not an error. REQ-2.3.5
    return {month: tables[month] for month in months if month in tables}


# --- Diff engine ------------------------------------------------------------------------


@dataclass
class ChangeRow:
    """One output row: a key, what happened to it, and the data that goes with it."""

    month: str
    key: Any
    status: str
    values: Dict[str, Any]

    def as_dict(self, key_column: str, columns: Sequence[str]) -> Dict[str, Any]:
        """Flatten to the output shape: Month, <key>, Change_Status, then the rest. REQ-3.7.1"""
        row: Dict[str, Any] = {
            MONTH_COLUMN: self.month,
            key_column: self.key,
            STATUS_COLUMN: self.status,
        }
        for column in columns:
            row[column] = self.values.get(column)  # REQ-3.7.3 missing column -> None
        return row


@dataclass
class DiffResult:
    """The full diff: the output column order plus the rows, ready to write."""

    key_column: str
    columns: List[str]
    changes: List[ChangeRow]

    @property
    def header(self) -> List[str]:
        return [MONTH_COLUMN, self.key_column, STATUS_COLUMN, *self.columns]

    def as_dicts(self) -> List[Dict[str, Any]]:
        return [change.as_dict(self.key_column, self.columns) for change in self.changes]

    def __len__(self) -> int:
        return len(self.changes)


def _sort_key(value: Any) -> "tuple[int, float, str]":
    """Numbers first and numerically, then text case-insensitively, then nulls last. REQ-3.6.2

    Excel hands us ints, floats, strings and blanks in one column and those are not mutually
    comparable in Python 3 -- hence the explicit rank.
    """
    if value is None:
        return (2, 0.0, "")
    if isinstance(value, bool):
        return (1, 0.0, str(value).casefold())
    if isinstance(value, (int, float)):
        return (0, float(value), "")
    text = str(value)
    try:
        return (0, float(text), "")
    except ValueError:
        return (1, 0.0, text.casefold())


def _values_differ(
    previous: Mapping[str, Any],
    current: Mapping[str, Any],
    columns: Sequence[str],
    key: str,
) -> bool:
    """Compare only the current month's non-key columns. REQ-3.2.5"""
    return any(
        previous.get(column) != current.get(column) for column in columns if column != key
    )


def compare_months(
    previous: Optional[MonthTable],
    current: MonthTable,
    key: str = DEFAULT_KEY,
    detect_modified: bool = False,
) -> List[ChangeRow]:
    """Compare one month against its predecessor. ``previous=None`` means base month."""
    if previous is None:  # REQ-3.4.1
        return [
            ChangeRow(current.month, row[key], BASE_MONTH, row)
            for row in current.rows
            if row.get(key) is not None  # REQ-3.4.2
        ]

    current_rows = current.by_key(key)    # REQ-3.1.1 full outer join on the key
    previous_rows = previous.by_key(key)

    changes: List[ChangeRow] = []

    for row_key, row in current_rows.items():
        if row_key not in previous_rows:
            changes.append(ChangeRow(current.month, row_key, ADDED, row))  # REQ-3.2.1 / REQ-3.3.1
            continue

        status = UNCHANGED  # REQ-3.2.3
        if detect_modified and _values_differ(previous_rows[row_key], row, current.columns, key):
            status = MODIFIED  # REQ-3.2.4
        changes.append(ChangeRow(current.month, row_key, status, row))

    # A Removed row carries the PREVIOUS month's values (the month it was last seen in,
    # REQ-3.3.2) but is labeled with the CURRENT month (where the removal was observed,
    # REQ-3.3.3). That asymmetry is the whole point of the "with_name" variant.
    for row_key, row in previous_rows.items():
        if row_key not in current_rows:
            changes.append(ChangeRow(current.month, row_key, REMOVED, row))  # REQ-3.2.2

    changes.sort(key=lambda change: _sort_key(change.key))
    return changes


def diff_tables(
    tables: Mapping[str, MonthTable],
    key: str = DEFAULT_KEY,
    months: Optional[Sequence[str]] = None,
    include_unchanged: bool = False,
    detect_modified: bool = False,
) -> DiffResult:
    """Walk the months in order and collect every change. REQ-1.2 / REQ-3.6.1

    A month whose immediate predecessor is absent from the workbook is treated as a base
    month, exactly as ``List.Generate`` does in the M template: it carries ``Prev`` forward
    as null and the comparison short-circuits rather than reaching further back. REQ-3.5.1
    """
    months = list(months) if months else MONTHS

    changes: List[ChangeRow] = []
    columns: List[str] = []
    seen_columns = set()

    for index, month in enumerate(months):
        current = tables.get(month)
        if current is None:
            continue

        previous = tables.get(months[index - 1]) if index > 0 else None
        changes.extend(compare_months(previous, current, key, detect_modified))

        # The union of every month's columns, in first-seen order. REQ-3.7.2
        for column in current.columns:
            if column != key and column not in seen_columns:
                seen_columns.add(column)
                columns.append(column)

    if not include_unchanged:  # REQ-3.8.1
        changes = [change for change in changes if change.status in CHANGE_STATUSES]

    return DiffResult(key_column=key, columns=columns, changes=changes)


# --- Writer -----------------------------------------------------------------------------


def format_for(path: "str | Path | None", requested: "str | None") -> str:
    """Explicit --format wins, else infer from the extension, else 'table'. REQ-4.2"""
    if requested:
        return requested
    if path is not None:
        suffix = Path(path).suffix.casefold()
        if suffix in (".xlsx", ".xlsm"):
            return "xlsx"
        if suffix == ".json":
            return "json"
        if suffix in (".csv", ".txt"):
            return "csv"
    return "table"


def _scalar(value: Any) -> Any:
    """Excel dates arrive as datetimes; make them JSON- and CSV-safe. REQ-4.6"""
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def _text(value: Any) -> str:
    return "" if value is None else str(_scalar(value))  # REQ-4.7


def to_table(result: DiffResult) -> str:
    """A fixed-width table for the terminal. REQ-4.3"""
    header = result.header
    rows = [[_text(row[column]) for column in header] for row in result.as_dicts()]

    widths = [len(column) for column in header]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))

    def line(cells: List[str]) -> str:
        return "  ".join(cell.ljust(widths[i]) for i, cell in enumerate(cells)).rstrip()

    out = [line(header), "  ".join("-" * width for width in widths)]
    out.extend(line(row) for row in rows)
    return "\n".join(out)


def to_csv(result: DiffResult) -> str:
    """REQ-4.4"""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=result.header, extrasaction="ignore")
    writer.writeheader()
    for row in result.as_dicts():
        writer.writerow({column: _scalar(value) for column, value in row.items()})
    return buffer.getvalue()


def to_json(result: DiffResult) -> str:
    """REQ-4.5"""
    payload = [
        {column: _scalar(value) for column, value in row.items()} for row in result.as_dicts()
    ]
    return json.dumps(payload, indent=2, default=str)


def write_xlsx(result: DiffResult, path: "str | Path") -> None:
    """Write the changes to a single styled sheet named 'Changes'. REQ-4.1.1"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Changes"

    header = result.header
    rows = result.as_dicts()

    worksheet.append(header)
    for row in rows:
        worksheet.append([_scalar(row[column]) for column in header])

    status_index = header.index(STATUS_COLUMN)

    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        fill = _STATUS_FILLS.get(row[status_index].value)
        if fill is not None:
            for cell in row:
                cell.fill = fill

    for index, column in enumerate(header, start=1):  # REQ-4.1.4
        longest = max([len(column)] + [len(_text(row[column])) for row in rows])
        worksheet.column_dimensions[get_column_letter(index)].width = min(longest + 4, 40)

    # An Excel Table needs at least one data row; below that, fall back to an autofilter so a
    # zero-change run still produces a valid workbook. REQ-4.1.6
    span = f"A1:{get_column_letter(len(header))}{worksheet.max_row}"
    if worksheet.max_row > 1:
        table = Table(displayName="tbl_Changes", ref=span)  # REQ-4.1.5
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = span

    worksheet.freeze_panes = "A2"  # REQ-4.1.7
    workbook.save(Path(path))


def render(result: DiffResult, fmt: str) -> str:
    """Render to a string. ``xlsx`` is binary and must go through write_xlsx instead. REQ-4.8"""
    if fmt == "table":
        return to_table(result)
    if fmt == "csv":
        return to_csv(result)
    if fmt == "json":
        return to_json(result)
    raise ValueError(f"{fmt!r} cannot be rendered as text; use write_xlsx()")


# --- CLI --------------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="xldiff_core",
        description=(
            "Compare the monthly tables in an Excel workbook and output only the rows that "
            "changed -- added and removed rows, with all their columns."
        ),
        epilog=(
            "The workbook should hold one table per month, either as Excel tables named "
            "tbl_Jan...tbl_Dec or as sheets named Jan...Dec. Each month is compared against "
            "the month before it."
        ),
    )
    parser.add_argument("workbook", type=Path, help="path to the .xlsx workbook to compare")
    parser.add_argument("-o", "--output", type=Path, help="write the result here (default: stdout)")
    parser.add_argument("-f", "--format", choices=FORMATS, help="output format (default: inferred)")
    parser.add_argument("-k", "--key", default=DEFAULT_KEY, metavar="COLUMN",
                        help=f"column identifying a row across months (default: {DEFAULT_KEY})")
    parser.add_argument("--table-prefix", default=DEFAULT_TABLE_PREFIX, metavar="PREFIX",
                        help=f"prefix of the per-month table names (default: {DEFAULT_TABLE_PREFIX})")
    parser.add_argument("--months", metavar="LIST",
                        help="comma-separated months, in order, to compare (default: Jan..Dec)")
    parser.add_argument("--detect-modified", action="store_true",
                        help="also report same-key/changed-value rows as 'Modified'")
    parser.add_argument("--all", action="store_true",
                        help="include Unchanged and Base Month rows too")
    return parser


def parse_months(raw: Optional[str]) -> List[str]:
    """REQ-5.4"""
    if not raw:
        return list(MONTHS)

    months = [month.strip() for month in raw.split(",") if month.strip()]
    unknown = [month for month in months if month not in MONTHS]
    if unknown:
        raise WorkbookError(
            f"Unknown month(s): {', '.join(unknown)}. Expected any of: {', '.join(MONTHS)}."
        )
    if len(months) < 2:
        raise WorkbookError("--months needs at least two months to have anything to compare.")
    return months


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)

    try:
        months = parse_months(args.months)
        tables = read_workbook(
            args.workbook, key=args.key, table_prefix=args.table_prefix, months=months
        )
        result = diff_tables(
            tables,
            key=args.key,
            months=months,
            include_unchanged=args.all,
            detect_modified=args.detect_modified,
        )
    except WorkbookError as exc:
        print(f"xldiff: {exc}", file=sys.stderr)  # REQ-5.9
        return EXIT_ERROR

    fmt = format_for(args.output, args.format)

    if fmt == "xlsx" and args.output is None:  # REQ-5.5
        print("xldiff: --format xlsx needs an --output path to write to.", file=sys.stderr)
        return EXIT_ERROR

    if args.output is None:
        print(render(result, fmt))  # REQ-5.6
    else:
        args.output.parent.mkdir(parents=True, exist_ok=True)  # REQ-5.7
        if fmt == "xlsx":
            write_xlsx(result, args.output)
        else:
            args.output.write_text(render(result, fmt), encoding="utf-8")

        found = ", ".join(sorted(tables, key=months.index))
        print(  # REQ-5.8 summary goes to stderr, keeping stdout clean for the data
            f"xldiff: {len(result)} changed row(s) across {found} -> {args.output}",
            file=sys.stderr,
        )

    return EXIT_OK  # REQ-5.10


if __name__ == "__main__":
    raise SystemExit(main())
